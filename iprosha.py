import logging
import argparse
import json
from collections import namedtuple
import re
import pprint
import pickle

import requests
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

logging.basicConfig(level=logging.INFO)
logging.getLogger("requests").setLevel(logging.WARNING)
logger = logging.getLogger()


def make_cmd_arguments_parser():
    parser = argparse.ArgumentParser()
    parser.add_argument('--session', '-s', type=int, help='session-id auth method don\'t need sms. Paste session number value')
    return parser


def get_ipro_auth_session(login, password, session_id=None):
    session = requests.Session()
    UserInfo = namedtuple('AuthInfo', ['login', 'password', 'man_id', 'login_type', 'session_id'])
    man_id, login_type = '9019820', 'WI'
    login_with_sms_url = 'https://spb.etm.ru/ns2000/data-login.php'
    login_without_sms_url = 'http://spb.etm.ru/ipro2'
    if session_id:
        request_params = {
            'login': login,
            'session': session_id,
            'man-id': man_id,
            'login_type': login_type
        }
        session.get(login_without_sms_url, params=request_params)
        logger.info('You authorise with exist session-id: {}'.format(session_id))
    else:
        request_params = {
            't': 'login',
            'log': login,
            'pwd': password
        }
        session_params_json = session.get(login_with_sms_url, params=request_params).json()
        session_id = session_params_json.get('session')
        man_id = session_params_json.get('man_id')
        login_type = session_params_json.get('login_type')
        sms_code = input('Please type sms code here: ')
        request_params['smsCode'], request_params['session'] = sms_code, session_id
        session.get(login_with_sms_url, params=request_params)

    logger.info('You authorise! You man-id: {}'.format(man_id))
    logger.info(' You session-id: {}'.format(session_id))
    logger.info(' You login-type: {}'.format(login_type))
    session_info = UserInfo(login=login,
                            man_id=man_id,
                            password=password,
                            login_type=login_type,
                            session_id=session_id)
    return session, session_info


def clean_html_tags(cleaned_text):
    return BeautifulSoup(cleaned_text, "html.parser").text


def fetch_full_client_list(session, session_info, url_id_parameter):
    client_info_page_url = 'http://spb.etm.ru/cat/data-orgtree.html'
    client_info_page_params = {
        'login': session_info.login,
        'man': session_info.man_id,
        'id': url_id_parameter,
        'd1': '01/09/17',
        'd2': '01/10/17',
    }
    client_info_page = session.get(client_info_page_url, params=client_info_page_params)
    client_info_json = json.loads(client_info_page.text)['Nodes']
    ClientInfo = namedtuple('ClientInfo', ['id', 'name', 'specialization', 'worth', 'warm'])
    client_info_list = []
    for client_info in client_info_json:
        client_id = client_info['ID']
        name, specialization, worth, warm = re.match(r'(.+) \((\S+),~(\S+), +(.+)\)', client_info['Name']).groups()
        client_info_list.append(ClientInfo(
            id=client_id,
            name=clean_html_tags(name),
            specialization=specialization,
            worth=worth,
            warm=warm
        ))
    return client_info_list


def save_list_of_named_tuples_to_xlsx(list_of_named_tuples, output_filepath):
    if not list_of_named_tuples:
        raise ValueError('You need put some data to write')
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet()
    worksheet.append(list_of_named_tuples[0]._fields)
    for named_tuple in list_of_named_tuples:
        worksheet.append(named_tuple)
    workbook.save(output_filepath)


def fill_schedule_task(session, session_info, client_code, task_date, task_message, task_result=None):
    client_schedule_post_url = 'http://ipro.etm.ru/cat/runprog.html'
    schedule_request_params = {
        'man': session_info.man_id,
        'login': session_info.login,
        'pme_persons': 'pmp_class37^ССПб3$man-code^{man_id}$cli-code^{client_code}$exm_mancode^'.format(man_id=session_info.man_id,
                                                                                                        client_code=client_code),
        'pme_datep': task_date,
        'pme_task': task_message,
        'pme_result': task_result,
        # Unknown params
        'syf_prog': 'pr_meeting-rsp',
        'withoutArchive': 'yes',
        'RSPAction': 'A',
        'pme_state': 'appoint',
        'RO_theme': 'Развитие продаж',
        'pme_theme': 'ВТ10',
        'RO_subtheme': 'Встречи с партнёрами',
        'pme_subtheme': 'ВТ1010',
        'RO_type': 'Встреча с партнёром',
        'pme_type': 'ВМ10',
    }
    session.get(client_schedule_post_url, params=schedule_request_params)


def fill_schedule(session, session_info, schedule_file): # $("#dialog_delete").dialog("open");
    workbook = load_workbook(filename=schedule_file)
    worksheet = workbook.worksheets[0]
    schedule_file_data = []
    for row in worksheet.rows:
        schedule_file_data.append([cell.value for cell in row])

    for task in schedule_file_data[1:]:
        client_code = task[0]
        task_date = task[5]
        task_message = task[6]
        fill_schedule_task(session=session,
                           session_info=session_info,
                           client_code=client_code,
                           task_date=task_date,
                           task_message=task_message,
                           task_result=None)


# def save_auth_info_to_pickle():
#     with open('auth.pk', 'rb') as file_handler:
#         auth_info_data = pickle.load(file_handler)
#     login = 'test1'
#     current_auth_info = {
#         'password': 11,
#         'man_id': 44,
#         'login_type': 11,
#         'session_id': 11
#     }
#     auth_info_data[login] = current_auth_info
#     with open('auth.pk', 'wb') as file_handler:
#         pickle.dump(auth_info_data, file_handler)


if __name__ == '__main__':
    login, password = '07elv', 'd561tj4'
    cmd_args_parser = make_cmd_arguments_parser()
    cmd_args = cmd_args_parser.parse_args()
    session_id = cmd_args.session

    ipro_session, ipro_auth_info = get_ipro_auth_session(login=login,
                                                         password=password,
                                                         session_id=session_id,)

    # pprint.pprint(ipro_auth_info)

    # full_client_list = fetch_full_client_list(session=ipro_session,
    #                                           session_info=ipro_auth_info,
    #                                           url_id_parameter=1461216007)
    # save_list_of_named_tuples_to_xlsx(list_of_named_tuples=full_client_list,
    #                                   output_filepath='full_client_list_{}.xlsx'.format(login))
    # fill_schedule(session=ipro_session,
    #               session_info=ipro_auth_info,
    #               schedule_file='full_client_list_07elv.xlsx',)