import logging
import json
from collections import namedtuple
import re
import pprint
import pickle

import requests
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

from .db import AuthDbAbstract

logging.basicConfig(level=logging.INFO)
logging.getLogger("requests").setLevel(logging.WARNING)
logger = logging.getLogger()


class IProAuthSession:
    IPRO_URI = 'http://spb.etm.ru/ipro2'
    AUTH_URI = 'https://spb.etm.ru/ns2000/data-login.php'
    
    def __init__(self, login, auth_db):
        self._login = login
        self._man_id=None
        self._password=None
        self._login_type='WI'
        self._session_id=None
        
        self._session = requests.Session()                    
        self._db = auth_db
        self._is_login_data_in_db = False
        if self._db.is_exist_by_login(self._login):
            login_db_data = self._db.get_by_login(self._login)
            self._man_id = login_db_data.get('man_id')
            self._password = login_db_data.get('password')
            self._last_session_id = login_db_data.get('last_session_id')
            self._is_login_data_in_db = True
            
    def prepare_login_with_sms(self, password):
        self._password = password
        request_params = {
            't': 'login',
            'log': self._login,
            'pwd': self._password,
        }
        auth_json_response = self._session.get(self.AUTH_URI, params=request_params).json()
        if auth_json_response['msg'] == 'Неверный логин или пароль!':
            raise PermissionError('Wrong login or password')
        logger.info(auth_json_response)
        self._session_id = auth_json_response.get('session')
        self._man_id = auth_json_response.get('man_id')
        self._login_type = auth_json_response.get('login_type')
        
        self._db.set_by_login(
            self._login,
            login=self._login,
            password=self._password,
            man_id=self._man_id,
            login_type=self._login_type,
            last_session_id=self._session_id
        )
        logger.info('Auth session prepared')
        
        
    def login_with_sms(self, sms_code):
        if not all([self._man_id, self._session_id]):
            raise AttributeError('Call prepare_login_with_sms before or initialize DB data to login:{}'.format(self._login))
        request_params = {
            't': 'login',
            'log': self._login,
            'pwd': self._password,
            'smsCode': sms_code,
            'session': self._session_id,
        }
        logger.info('Fetching auth session for {}...'.format(self._login))
        self._session.get(self.AUTH_URI, params=request_params)
        logger.info('Auth session saved')
        seamless_auth_url = '{url}/?login={login}&session={session_id}&man-id={man_id}&login_type={login_type}'.format(
            url=self.IPRO_URI,
            login=self._login,
            session_id=self._session_id,
            man_id=self._man_id,
            login_type=self._login_type,
            )
        logger.info('Use {} to seamless auth'.format(seamless_auth_url))
        
    def login_with_session_id(self, session_id):
        pass
    
    @property
    def login(self):
        return self._login
    
    @property
    def password(self):
        return self._password
        
    @property
    def man_id(self):
        return self._man_id
        
    @property
    def login_type(self):
        return self._login_type
        
    @property
    def session_id(self):
        return self._session_id
        
    @property
    def session(self):
        return self._session

    

    
    

# def get_ipro_auth_session(login, password, session_id=None):
#     session = requests.Session()
#     UserInfo = namedtuple('AuthInfo', ['login', 'password', 'man_id', 'login_type', 'session_id'])
#     man_id, login_type = '9019820', 'WI'
#     login_with_sms_url = 'https://spb.etm.ru/ns2000/data-login.php'
#     login_without_sms_url = 'http://spb.etm.ru/ipro2'
#     if session_id:
#         request_params = {
#             'login': login,
#             'session': session_id,
#             'man-id': man_id,
#             'login_type': login_type
#         }
#         session.get(login_without_sms_url, params=request_params)
#         logger.info('You authorise with exist session-id: {}'.format(session_id))
#     else:
#         request_params = {
#             't': 'login',
#             'log': login,
#             'pwd': password
#         }
#         session_params_json = session.get(login_with_sms_url, params=request_params).json()
#         session_id = session_params_json.get('session')
#         man_id = session_params_json.get('man_id')
#         login_type = session_params_json.get('login_type')
#         sms_code = input('Please type sms code here: ')
#         request_params['smsCode'], request_params['session'] = sms_code, session_id
#         session.get(login_with_sms_url, params=request_params)

#     logger.info('You authorise! You man-id: {}'.format(man_id))
#     logger.info(' You session-id: {}'.format(session_id))
#     logger.info(' You login-type: {}'.format(login_type))
#     session_info = UserInfo(login=login,
#                             man_id=man_id,
#                             password=password,
#                             login_type=login_type,
#                             session_id=session_id)
#     return session, session_info


# def clean_html_tags(cleaned_text):
#     return BeautifulSoup(cleaned_text, "html.parser").text


# def fetch_full_client_list(session, session_info, url_id_parameter):
#     client_info_page_url = 'http://spb.etm.ru/cat/data-orgtree.html'
#     client_info_page_params = {
#         'login': session_info.login,
#         'man': session_info.man_id,
#         'id': url_id_parameter,
#         'd1': '01/10/17',
#         'd2': '01/11/17',
#     }
#     client_info_page = session.get(client_info_page_url, params=client_info_page_params)
#     client_info_json = json.loads(client_info_page.text)['Nodes']
#     ClientInfo = namedtuple('ClientInfo', ['id', 'name', 'specialization', 'worth', 'warm'])
#     client_info_list = []
#     for client_info in client_info_json:
#         client_id = client_info['ID']
#         name, specialization, worth, warm = re.match(r'(.+) \((\S+),~(\S+), +(.+)\)', client_info['Name']).groups()
#         client_info_list.append(ClientInfo(
#             id=client_id,
#             name=clean_html_tags(name),
#             specialization=specialization,
#             worth=worth,
#             warm=warm
#         ))
#     return client_info_list


# def save_list_of_named_tuples_to_xlsx(list_of_named_tuples, output_filepath):
#     if not list_of_named_tuples:
#         raise ValueError('You need put some data to write')
#     workbook = Workbook(write_only=True)
#     worksheet = workbook.create_sheet()
#     worksheet.append(list_of_named_tuples[0]._fields)
#     for named_tuple in list_of_named_tuples:
#         worksheet.append(named_tuple)
#     workbook.save(output_filepath)


# def setup_plan_to_company(session, session_info, company_id, to_for_month, rtn_for_month):
#     setup_plan_url = 'http://spb.etm.ru/cat/data-plan.html'
#     setup_plan_params = {
#         'id' : company_id,
#         'login': session_info.login,
#         'man' : session_info.man_id,
#         'usr': session_info.man_id,
#         'month': '11',
#         'year': '2017',
#         'to': to_for_month,
#         'rtn': rtn_for_month,
#         'cause': 'завершение объекта',
#     }
#     session.post(setup_plan_url, params=setup_plan_params)


# def fill_schedule_task(session, session_info, client_code, task_date, task_message, task_result=None):
#     client_schedule_post_url = 'http://ipro.etm.ru/cat/runprog.html'
#     schedule_request_params = {
#         'man': session_info.man_id,
#         'login': session_info.login,
#         'pme_persons': 'pmp_class37^ССПб3$man-code^{man_id}$cli-code^{client_code}$exm_mancode^'.format(man_id=session_info.man_id,
#                                                                                                         client_code=client_code),
#         'pme_datep': task_date,
#         'pme_task': task_message,
#         'pme_result': task_result,
#         # Unknown params
#         'syf_prog': 'pr_meeting-rsp',
#         'withoutArchive': 'yes',
#         'RSPAction': 'A',
#         'pme_state': 'appoint',
#         'RO_theme': 'Развитие продаж',
#         'pme_theme': 'ВТ10',
#         'RO_subtheme': 'Встречи с партнёрами',
#         'pme_subtheme': 'ВТ1010',
#         'RO_type': 'Встреча с партнёром',
#         'pme_type': 'ВМ10',
#     }
#     session.get(client_schedule_post_url, params=schedule_request_params)


# def fill_schedule(session, session_info, schedule_file): # $("#dialog_delete").dialog("open");
#     workbook = load_workbook(filename=schedule_file)
#     worksheet = workbook.worksheets[0]
#     schedule_file_data = []
#     for row in worksheet.rows:
#         schedule_file_data.append([cell.value for cell in row])

#     for task in schedule_file_data[1:]:
#         client_code = task[0]
#         task_date = task[5]
#         task_message = task[6]
#         fill_schedule_task(session=session,
#                           session_info=session_info,
#                           client_code=client_code,
#                           task_date=task_date,
#                           task_message=task_message,
#                           task_result=None)
