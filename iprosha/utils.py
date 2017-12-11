from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from collections import namedtuple


# TODO: проверить где используется эта функция
def clean_html_tags(cleaned_text):
    return BeautifulSoup(cleaned_text, "html.parser").text


def save_list_of_named_tuples_to_xlsx(list_of_named_tuples, output_filepath):
    if not list_of_named_tuples:
        raise ValueError('You need put some data to write')
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet()
    worksheet.append(list_of_named_tuples[0]._fields)
    for named_tuple in list_of_named_tuples:
        worksheet.append(named_tuple)
    workbook.save(output_filepath)


def add_new_headers_to_xlsx(xlsx_filepath, list_of_added_headers):
    workbook = load_workbook(filename=xlsx_filepath)
    worksheet = workbook.worksheets[0]
    for index, value in enumerate(list_of_added_headers, start=worksheet.max_column+1):
        worksheet.cell(row=1, column=index, value=value)
    workbook.save(xlsx_filepath)


def create_list_of_named_tuples_from_xslx(xlsx_filepath):
    workbook = load_workbook(filename=xlsx_filepath)
    worksheet = workbook.worksheets[0]
    header = []
    for row in worksheet.iter_rows(max_row=1):
        for cell in row:
            header.append(cell.value)
    ExcelRow = namedtuple('ExcelRow', header)
    worksheet_data_list = []
    for row in worksheet.iter_rows(min_row=2):
        worksheet_data_list.append(ExcelRow._make([cell.value for cell in row]))
    return worksheet_data_list
